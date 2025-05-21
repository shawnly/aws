import boto3
import pandas as pd
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def parse_arguments():
    parser = argparse.ArgumentParser(description='Get AWS Cost Explorer data for a specific account')
    parser.add_argument('--start-date', required=True, help='Start date in YYYY-MM-DD format')
    parser.add_argument('--end-date', required=True, help='End date in YYYY-MM-DD format')
    parser.add_argument('--account-id', required=True, help='AWS Account ID')
    return parser.parse_args()

def validate_dates(start_date, end_date):
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        raise ValueError("Dates must be in YYYY-MM-DD format")
    
    if start_date > end_date:
        raise ValueError("Start date must be before end date")

def get_cost_and_usage(client, start_date, end_date, account_id):
    response = client.get_cost_and_usage(
        TimePeriod={
            'Start': start_date,
            'End': end_date
        },
        Granularity='MONTHLY',
        Metrics=['UnblendedCost', 'UsageQuantity'],
        GroupBy=[
            {
                'Type': 'DIMENSION',
                'Key': 'SERVICE'
            },
            {
                'Type': 'DIMENSION',
                'Key': 'RECORD_TYPE'
            }
        ],
        Filter={
            'Dimensions': {
                'Key': 'LINKED_ACCOUNT',
                'Values': [account_id]
            }
        }
    )
    return response

def process_cost_data(response):
    data = []
    
    for result in response['ResultsByTime']:
        period_start = result['TimePeriod']['Start']
        period_end = result['TimePeriod']['End']
        
        for group in result['Groups']:
            service = group['Keys'][0]
            record_type = group['Keys'][1]
            cost = float(group['Metrics']['UnblendedCost']['Amount'])
            unit = group['Metrics']['UnblendedCost']['Unit']
            usage = float(group['Metrics']['UsageQuantity']['Amount'])
            
            data.append({
                'Period Start': period_start,
                'Period End': period_end,
                'Service': service,
                'Record Type': record_type,
                'Cost': cost,
                'Unit': unit,
                'Usage Quantity': usage
            })
    
    return pd.DataFrame(data)

def calculate_summary(df):
    # Group by period and service, calculating net cost (usage - refund)
    summary = df.pivot_table(
        index=['Period Start', 'Service'], 
        columns='Record Type', 
        values='Cost', 
        aggfunc='sum'
    ).fillna(0)
    
    # Add a Net column (Usage - Refund)
    if 'Refund' in summary.columns:
        summary['Net Cost'] = summary['Usage'] + summary['Refund']  # Refund amounts are negative
    else:
        summary['Net Cost'] = summary['Usage']
        
    # Reset index for easier manipulation
    summary = summary.reset_index()
    
    return summary

def save_to_excel(raw_data, summary_data, account_id, start_date, end_date):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"aws_cost_{account_id}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Save raw data
        raw_data.to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Save summary data
        summary_data.to_excel(writer, sheet_name='Cost Summary', index=False)
        
        # Format the worksheets
        workbook = writer.book
        
        # Format summary sheet
        summary_sheet = workbook['Cost Summary']
        summary_sheet.column_dimensions['A'].width = 12  # Period Start
        summary_sheet.column_dimensions['B'].width = 30  # Service
        
        # Add header with parameters
        summary_sheet.insert_rows(0, 3)
        summary_sheet.cell(row=1, column=1, value=f"AWS Cost Report for Account: {account_id}")
        summary_sheet.cell(row=2, column=1, value=f"Period: {start_date} to {end_date}")
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Highlight the Net Cost column header
        for col in range(1, summary_sheet.max_column + 1):
            if summary_sheet.cell(row=4, column=col).value == "Net Cost":
                summary_sheet.cell(row=4, column=col).font = Font(bold=True)
                summary_sheet.cell(row=4, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    return filename

def main():
    args = parse_arguments()
    validate_dates(args.start_date, args.end_date)
    
    try:
        # Initialize Cost Explorer client
        ce_client = boto3.client('ce')
        
        print(f"Fetching cost data for account {args.account_id} from {args.start_date} to {args.end_date}...")
        response = get_cost_and_usage(ce_client, args.start_date, args.end_date, args.account_id)
        
        # Process raw data
        raw_data = process_cost_data(response)
        
        # Calculate summary
        summary_data = calculate_summary(raw_data)
        
        # Save to Excel
        output_file = save_to_excel(raw_data, summary_data, args.account_id, args.start_date, args.end_date)
        print(f"Cost data saved to {output_file}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()