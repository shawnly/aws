import boto3
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

def get_cost_data(start_date, end_date, account_id):
    client = boto3.client('ce')

    # Group by service
    results = client.get_cost_and_usage(
        TimePeriod={
            'Start': start_date,
            'End': end_date
        },
        Granularity='MONTHLY',
        Metrics=['UnblendedCost'],
        GroupBy=[
            {'Type': 'DIMENSION', 'Key': 'SERVICE'},
            {'Type': 'DIMENSION', 'Key': 'LINKED_ACCOUNT'}
        ],
        Filter={
            'Dimensions': {
                'Key': 'LINKED_ACCOUNT',
                'Values': [account_id]
            }
        }
    )
    return results['ResultsByTime']

def extract_and_write_results(raw_data, start_date, end_date, account_id):
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "RawCosts"
    ws_summary = wb.create_sheet(title="Summary")

    # Headers
    ws_raw.append(['Start', 'End', 'Account ID', 'Service', 'Amount ($)', 'Unit'])
    account_summary = defaultdict(float)

    for month_data in raw_data:
        start = month_data['TimePeriod']['Start']
        end = month_data['TimePeriod']['End']
        for group in month_data['Groups']:
            service = group['Keys'][0]
            acc = group['Keys'][1]
            amount = float(group['Metrics']['UnblendedCost']['Amount'])
            unit = group['Metrics']['UnblendedCost']['Unit']

            # Append to raw sheet
            ws_raw.append([start, end, acc, service, amount, unit])

            # Track totals
            account_summary[acc] += amount

    # Summary sheet
    ws_summary.append(['Account ID', 'Total Cost After Refund ($)'])
    for acc, total in account_summary.items():
        ws_summary.append([acc, round(total, 2)])

    # Autosize columns
    for sheet in [ws_raw, ws_summary]:
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2

    # Save file
    filename = f"aws_cost_report_{account_id}_{start_date}_to_{end_date}.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")

def main(start_date, end_date, account_id):
    print(f"Fetching cost data for account: {account_id} from {start_date} to {end_date}")
    raw_data = get_cost_data(start_date, end_date, account_id)
    extract_and_write_results(raw_data, start_date, end_date, account_id)

# Example usage
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument('--start-date', required=True, help='Start date in YYYY-MM-DD')
    parser.add_argument('--end-date', required=True, help='End date in YYYY-MM-DD')
    parser.add_argument('--account-id', required=True, help='12-digit AWS account ID')

    args = parser.parse_args()
    main(args.start_date, args.end_date, args.account_id)
